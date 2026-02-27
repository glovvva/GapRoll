"""Generate corrupted Excel files simulating Polish ERP exports."""
import argparse
import random
from openpyxl import Workbook
from openpyxl.styles import Font

try:
    from faker import Faker
    fake = Faker("pl_PL")
except Exception:
    fake = None


def _fake_first_name():
    if fake:
        return fake.first_name()
    return random.choice(["Anna", "Jan", "Maria", "Piotr", "Katarzyna", "Tomasz", "Magdalena", "Michał"])


def _fake_last_name():
    if fake:
        return fake.last_name()
    return random.choice(["Kowalski", "Nowak", "Wiśniewski", "Lewandowski", "Kamiński", "Zielińska"])


def _fake_pesel():
    if fake:
        try:
            return fake.pesel()
        except Exception:
            pass
    # Fallback: 11 digits, valid format
    year = random.randint(50, 99)
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    rest = "".join(str(random.randint(0, 9)) for _ in range(5))
    return f"{year:02d}{month:02d}{day:02d}{rest}"


def generate_dirty_excel(employees: int, corruption_rate: float, bugs: list, output: str):
    """Generate Excel with intentional data quality issues."""
    wb = Workbook()
    ws = wb.active

    # Company logo row (merged)
    ws.merge_cells("A1:F1")
    ws["A1"] = "FIRMA ABC Sp. z o.o. - Raport Płac"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = "Raport Płac - Styczeń 2026"
    # Header row 4
    ws["A4"] = "Imię"
    ws["B4"] = "Nazwisko"
    ws["C4"] = "PESEL"
    ws["D4"] = "Stanowisko"
    ws["E4"] = "Wynagrodzenie brutto"
    ws["F4"] = "Płeć"

    positions = ["Programista", "HR Manager", "Księgowa", "Dyrektor", "Specjalista"]
    duplicate_pesel_val = "85010112345"

    for i in range(employees):
        row = i + 5
        name = _fake_first_name()
        surname = _fake_last_name()
        pesel = _fake_pesel()
        position = random.choice(positions)
        salary_raw = random.randint(5000, 15000)
        gender = random.choice(["K", "M"])

        # Apply bugs
        salary = salary_raw
        if "mixed_encoding" in bugs and random.random() < corruption_rate / 100:
            salary = f"{salary_raw:,}".replace(",", " ")
        if "duplicate_pesel" in bugs and i % 30 == 0:
            pesel = duplicate_pesel_val
        if "invalid_salary" in bugs and random.random() < corruption_rate / 100:
            salary = "osiem tysięcy"
        if "missing_gender" in bugs and random.random() < 0.10:
            gender = ""

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=surname)
        ws.cell(row=row, column=3, value=pesel)
        ws.cell(row=row, column=4, value=position)
        ws.cell(row=row, column=5, value=salary)
        ws.cell(row=row, column=6, value=gender)

    # Apply merged cells after writing (so two rows share one salary cell)
    for i in range(employees):
        if "merged_cells" not in bugs or i % 20 != 0:
            continue
        row = i + 5
        if row + 1 <= employees + 4:
            ws.merge_cells(f"E{row}:E{row + 1}")

    wb.save(output)
    print(f"Generated {output}: {employees} employees, corruption ~{corruption_rate}%")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--employees", type=int, required=True)
    parser.add_argument("--corruption-rate", type=float, default=30)
    parser.add_argument("--bugs", type=str, required=True, help="Comma-separated: merged_cells,mixed_encoding,duplicate_pesel,invalid_salary,missing_gender")
    parser.add_argument("--output", type=str, required=True)
    args = parser.parse_args()
    bugs = [b.strip() for b in args.bugs.split(",")]
    generate_dirty_excel(args.employees, args.corruption_rate, bugs, args.output)


if __name__ == "__main__":
    main()
